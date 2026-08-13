# apps/base/reporting.py
"""
Exportación a Excel/PDF del listado de Tickets/Citas ya filtrado en pantalla
(Fase 15, sesión 21) — botón "Reporte" en los 3 paneles internos (Compras,
Vigilancia, Calidad) y en el Historial de Entregas del proveedor.

Vive en apps.base (no en apps.operations) por el mismo motivo que
apps.base.filters.resolver_periodo (Fase 10): lo consumen tanto
apps.operations (Ticket-based) como apps.appointments (Appointment-based,
donde no todas las citas tienen Ticket todavía) — evita el import cruzado
entre apps de negocio.

Un único formato de fila común (HistorialRow) permite reutilizar el mismo
exportador (exportar_excel/exportar_pdf) sin importar si el origen fue un
queryset de Ticket o de Appointment; ticket_a_row()/cita_a_row() son los
únicos puntos que conocen esa diferencia.
"""
from dataclasses import dataclass
from io import BytesIO

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

# Mismas columnas que _partials/historial_tickets.html (Ticket, OC(s),
# Proveedor, Fecha cita, Estado, Etapa activa) — sin la columna de acción
# "Trazabilidad" (no aplica a un archivo exportado) — más 'Sede' (sesión
# 48d), que la tabla en pantalla todavía no muestra (no pedido, ver
# CLAUDE.md) pero el archivo exportado sí.
COLUMNAS = ('Ticket', 'OC(s)', 'Proveedor', 'Sede', 'Fecha de cita', 'Estado', 'Etapa activa')

_ANCHOS_COLUMNA = (10, 24, 30, 20, 14, 16, 28)


@dataclass
class HistorialRow:
    ticket: str
    ocs: str
    proveedor: str
    sede: str
    fecha_cita: str
    estado: str
    etapa_activa: str

    def as_tuple(self):
        return (
            self.ticket, self.ocs, self.proveedor, self.sede,
            self.fecha_cita, self.estado, self.etapa_activa,
        )


def _etapa_activa_de_ticket(ticket) -> str:
    """Mismo criterio que _partials/historial_tickets.html: CANCELADO no
    tiene un valor propio en etapa_actual, se muestra como texto fijo."""
    return 'Cancelado' if ticket.estado == 'CANCELADO' else ticket.get_etapa_actual_display()


def ticket_a_row(ticket) -> HistorialRow:
    """Fila a partir de un Ticket (paneles Compras/Vigilancia/Calidad)."""
    ocs = ', '.join(f'OC {po.doc_num}' for po in ticket.appointment.purchase_orders.all())
    proveedor_user = ticket.appointment.user
    return HistorialRow(
        ticket=f'#{ticket.id}',
        ocs=ocs or '—',
        proveedor=proveedor_user.get_full_name() or proveedor_user.username,
        sede=ticket.appointment.sede.nombre,
        fecha_cita=ticket.appointment.slot.date.strftime('%d/%m/%Y') if ticket.appointment.slot else '—',
        estado=ticket.get_estado_display(),
        etapa_activa=_etapa_activa_de_ticket(ticket),
    )


def cita_a_row(cita) -> HistorialRow:
    """
    Fila a partir de un Appointment (Historial de Entregas del proveedor):
    no todas las citas tienen Ticket generado todavía (SOLICITADO/RECHAZADO),
    así que Ticket/Estado/Etapa activa caen a un fallback basado en la cita
    misma en ese caso, en vez de fallar.
    """
    ocs = ', '.join(f'OC {oc.doc_num}' for oc in cita.purchase_orders.all())
    ticket = getattr(cita, 'ticket', None)
    if ticket:
        ticket_label = f'#{ticket.id}'
        estado = ticket.get_estado_display()
        etapa_activa = _etapa_activa_de_ticket(ticket)
    else:
        ticket_label = '—'
        estado = cita.get_status_display()
        etapa_activa = '—'
    return HistorialRow(
        ticket=ticket_label,
        ocs=ocs or '—',
        proveedor=cita.user.get_full_name() or cita.user.username,
        sede=cita.sede.nombre,
        fecha_cita=cita.slot.date.strftime('%d/%m/%Y') if cita.slot else '—',
        estado=estado,
        etapa_activa=etapa_activa,
    )


def exportar_excel(rows, filename: str, titulo: str) -> HttpResponse:
    """Genera el .xlsx vía openpyxl y lo devuelve como descarga adjunta."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial'

    ws.append([titulo])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS))

    ws.append([f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")} · {len(rows)} registro(s)'])
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color='666666')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNAS))

    ws.append([])  # fila en blanco de separación

    header_row = ws.max_row + 1
    ws.append(COLUMNAS)
    header_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
    for col in range(1, len(COLUMNAS) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append(row.as_tuple())

    for i, ancho in enumerate(_ANCHOS_COLUMNA, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def exportar_pdf(rows, filename: str, titulo: str) -> HttpResponse:
    """
    Genera el .pdf vía xhtml2pdf, mismo enfoque de plantilla standalone
    (sin Bootstrap, CSS 2.1 simple) ya usado en cargo_ticket_pdf.html (Fase 12).
    """
    html = render_to_string('base/reporte_historial_pdf.html', {
        'titulo': titulo,
        'columnas': COLUMNAS,
        'filas': [row.as_tuple() for row in rows],
        'fecha_generacion': timezone.now(),
    })
    buffer = BytesIO()
    resultado = pisa.CreatePDF(html, dest=buffer)
    if resultado.err:
        return HttpResponse('No se pudo generar el PDF del reporte.', status=500)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response
